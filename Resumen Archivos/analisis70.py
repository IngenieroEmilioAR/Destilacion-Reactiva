from CodeLibrary import Simulation
import os

import numpy as np
from openpyxl import Workbook, load_workbook

wb = load_workbook("Resultados2.xlsx")
headers = ["Nstage", "fraccion_sep_spliiter", "etapa_alim_etanol", "etapa_alim_acido", "carga_termica_reherv", "Flujo_molar_domo","composicion_molar_acetato_domo", "convergence"]


ws = wb.create_sheet("70")
ws.append(headers)

fracciones = np.arange(0.3,0.8,0.2)

wd = os.getcwd()
BLOCKNAME = "COLREACT"


simu =Simulation(AspenFileName= "destilacion reactiva evaluacion70.bkp", WorkingDirectoryPath= wd ,VISIBILITY=False)

stages = 70
simu.BLK_RADFRAC_Set_NSTAGE(BLOCKNAME, stages)
stages_2 = np.arange(1,stages-1,6)
print("etapas", stages)
for fraccion in fracciones:
    simu.BLK_SPLITTER_Set_By_SplitFraction("B5", "REFLUJO", fraccion)
    for i in range(len(stages_2)):
        simu.BLK_RADFRAC_Set_FeedStage(BLOCKNAME, stages_2[i], "ALIMACID")
        stages_3 = np.arange(stages-1,stages_2[i],-6)

        for j in range(len(stages_3)):
            simu.BLK_RADFRAC_Set_FeedStage(BLOCKNAME, stages_3[j], "ALIMETOH")
            convergence = simu.Run()
                    
            heat_duty = simu.BLK_RADFRAC_Get_Reboiler_HeatDuty(BLOCKNAME)
            comp = simu.STRM_Get_MoleFracPerCompound("S8", "ACETATO")
            domo_flow = simu.STRM_Get_MoleFlowPerCompound("S8", "ACETATO")

            info = [stages, fraccion, stages_2[i], stages_3[j], heat_duty, domo_flow, comp, convergence]                    
            ws.append(info)
            wb.save("Resultados2.xlsx")

simu.CloseAspen()


print("Finalizado ", stages, " ETAPAS")